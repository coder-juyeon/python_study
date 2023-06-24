import requests
from bs4 import BeautifulSoup

# 액셀 import
import openpyxl

# 액셀의 스타일 import
from openpyxl.styles import Alignment, Font, colors, Border, Side, PatternFill

# 액셀 정렬 및 폰트
wrap_alignment = Alignment(wrap_text=True)
bold_font = Font(bold=True)
header_font = Font(bold=True, size = 14)

# 얇은 선 테두리 스타일
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
# 제목 행 색상
grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

# 액셀 기본 코드
excel_file = openpyxl.Workbook()
excel_sheet = excel_file.active

# active 시트 이름 변경
excel_sheet.title = '회사'

excel_sheet.append(['', '', '', ''])
# 액셀의 헤더 부분
header = ['No', '회사', '가격', '변동률']

# 헤더 append
excel_sheet.append([''] + header)

# 헤더에 스타일
for cell in excel_sheet[2][1:]:
    cell.font = header_font
    cell.border = thin_border
    cell.fill = grey_fill

excel_sheet.column_dimensions['A'].width = 5
excel_sheet.column_dimensions['B'].width = 10
excel_sheet.column_dimensions['C'].width = 35
excel_sheet.column_dimensions['D'].width = 20
excel_sheet.column_dimensions['E'].width = 20


res = requests.get('https://davelee-fun.github.io/blog/crawling_stock_example.html')
soup = BeautifulSoup(res.content, 'html.parser')

data = soup.select('li.row_sty')
index = 0
row_num = 2
for item in data:
    company = item.select_one('div.st_name').text.strip().replace("\n", "").replace(" ", "")
    price = item.select_one('div.st_price').text.strip().replace("\n", "").replace(" ", "")
    rate = item.select_one('div.st_rate').text.strip().replace("\n", "").replace(" ", "") 

    if company != None:
        index += 1
#         print(str(index)+". " + company+ " " + price + "원 " + rate)
        excel_data = [''] + [str(index)+". ", company, price + "원 ", rate]
        excel_sheet.append(excel_data)
        excel_sheet.cell(row = row_num, column = 2)
        
#         for cell in excel_sheet[row_num][1:]:
#             cell.alignment = wrap_alignment
#             cell.border = thin_border
        row_num += 1
        test = row_num
        for cell in excel_sheet[test][1:]:
            cell.border = thin_border

excel_file.save('company_info.xlsx')
excel_file.close()